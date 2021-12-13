#This function analyzes the cost of implement just a solar array with battery for the entire load


#COST OF SOLAR

import numpy as np
from numpy import matrix
from scipy.optimize import minimize
import math
import openpyxl




#CONSTNATS
LCOE_solar = 0.05           #$/kWh

#Battery Specs
battery_capacity = 13.5     #kWh
#battery_cost = 8500        #$
#battery_maintenance = 50   #$/year
battery_dc_voltage = 50     #V
battery_install_1 = 500     #$/kWh for a 2 hr system 
battery_install_2 = 25      #$/kWh for each additional hour of storage



#Solar Panel Specs
panel_cost = 220            #$/panel
panel_maintenance = 10      #$/year
panel_wattage = 0.425       #kW
efficiency = 0.90

#Inverter Specs
inverter_cost = 0           #$ #set to zero because battery: Tesla Powerwall 2 has inverter combined. 
inverter_maintenance = 0    #$/year

#Charge Controller Specs
controller_cost =15         #$/controller
controller_maintenance = 5  #$/year #accounts for having to replace controllers
controller_amperage = 30    #amps


#BUFFER
buffer = 1.2255 #Use this buffer to play around with what level of factor you want for the array
#buffer = 1
#peak load is 376.1 kW (from Nov 20th @ 19:00) / 306.875 (daily average from peak day) = 1.2255

#DATA
#Load data from January 30th 1990 (kW)
load_data = [278.4 ,276.2, 274.1, 271, 269.3, 279.7, 311.8, 299.3, 302.3, 313.2, 321.3, 322.5, 345, 336.7, 329.1, 325.5, 315.9, 318.3, 328.1, 364.6, 339.9, 315.6, 308.1, 288.1]
load_data = [element * buffer for element in load_data]

#Solar Data from June 9th 2014 (hourly average)
solar_data = [0, 0, 0, 0, 0 , 0.675, 2.883333333, 5.2, 6.966666667, 9.35, 10.775, 10.50833333, 9.366666667, 7.1, 3.95, 1.108333333, 0, 0, 0, 0, 0, 0, 0, 0] #June 9th

    
#Defining workbook we will save values to 
wb = openpyxl.Workbook()
sheet = wb.active




    
def cost_of_item(num_of_item, price_per_item):
    cost = num_of_item * price_per_item 
    return cost
    
       

def whole_number(battery_count):
    battery_count_int = battery_count.is_integer()
    if (battery_count_int) == False:
        battery_count += 1
        battery_count = math.trunc(battery_count)
    return battery_count



def inflation(present_value, rate, time):
    future_value = ((present_value) * ((1.0 + (rate))**(time)))
    return future_value



def build_spreadsheet():
    
    #Defining columns and titles
    sheet.cell(row = 1, column = 1).value = "Year"
    sheet.cell(row = 1, column = 2).value = "Cost"
    sheet.cell(row = 1, column = 3).value = "Total Cost"

    for j in range (1,22):
        sheet.cell(row = j+1, column = 1).value = (j-1)
    return


def avaliable_solar(): 

        
    lowest_value = 1000 #arbitary high starting point
    for j in range(135, 5, -1): #iterating to make solar data closer to load data
        solar_data_multiplied = [element * j for element in solar_data]
        
        if sum(solar_data_multiplied) >= sum(load_data):
            lowest_value = j
            lowest_solar_value = (solar_data_multiplied)
            life_energy = ((sum(lowest_solar_value)) * 365 * 20)

           
    
    #print(lowest_solar_value)
    print("SUMMARY:")
    print(sum(lowest_solar_value))
    print("Buffer: " + str(buffer))
    print("Total load with buffer to accomiate for is: " + str('{0:.2f}'.format(sum(load_data))) + (" kWh"))
    print("The multiplying factor of the solar data is " + str(lowest_value))
   # print("The new size of the solar array is: " + str(lowest_solar_value[10] * 100000) + (" kW"))
    print("The new size of the solar array is: " + str(lowest_value * 100) + (" kW"))
    print()
    
    #finding how much battery storage will be required
    solar_data_lowest_value = [element * lowest_value for element in solar_data]
    load_data_matrix = np.array(load_data)
    solar_data_matrix = np.array(solar_data_lowest_value)
    difference = load_data_matrix - solar_data_lowest_value


    print("QUANTITIES:")
    #Number of Solar Panels
    pv_array = ((lowest_solar_value[10])) #Maximum kW required to be output at the peak hour
    panel_count = pv_array / panel_wattage
    panel_count = whole_number(panel_count)
    print("Number of solar panels required: " + str('{0:.5g}'.format(panel_count)))
    
    #Finding battery storage
    battery_storage = 0
    battery_hours = 0
    for entry in difference:
        if entry < 0:
            battery_storage += entry
        if entry > 0:
            battery_hours += 1
            
            
    
    battery_storage = abs(battery_storage)
    print("Battery storage required: " + str('{0:.2f}'.format(battery_storage)) + (" kWh"))
    
    #DO NOT REALLY NEED THIS ONE BELOW
    #Number of batteries  
    battery_count = ((battery_storage / battery_capacity))
    battery_count = whole_number(battery_count)
    print("Number of batteries required: " + str(battery_count))
   
    #Number of Inverters
    #print("size inverter based off average hourly production : " + str((sum(load_data))/24))
    inverter_count = 0
    print("Number of inverters required: " + str(inverter_count))
    
    #Number of controllers
    controller_size = ((panel_count * (panel_wattage * 1000)) / battery_dc_voltage) #amp
    controller_count = controller_size / (controller_amperage)
    controller_count = whole_number(controller_count)
    print("Number of controllers required: " + str(controller_count))








   
    #cost of battery  
    #This is determining the cost of batteries
    battery_hour_load = battery_storage / battery_hours
    battery_initial_cost = 0 
  
    if battery_hours > 0:
        for k in range(battery_hours):   #this was an idea to step down battery hours until it's zero
            if k == 0:
                battery_initial_cost += battery_hour_load * 500   #cost to install batteries for the first 2 hours of a system
            if k == 1: 
                battery_initial_cost += battery_hour_load * 500   #cost to install batteries for the first 2 hours of a system
            if k >= 2:
                battery_initial_cost += battery_hour_load * 25   #cost to install batteries for the first 2 hours of a system

    print('{0:.2f}'.format(battery_initial_cost))
    print("range of battery_hours")
  
    
  
    
    #INITAL COSTS
    initial_cost_solar = cost_of_item(panel_count, panel_cost)
    initial_cost_inverter = cost_of_item(inverter_count, inverter_cost)
    initial_cost_controller = cost_of_item(controller_count, controller_cost) 
    
  
    
    #Printing findings
    print()
    print("COST")
    print("Cost of panels : $" +str(initial_cost_solar))
    print("Cost of battery: $" + str(battery_initial_cost))
    print("Cost of inverter: $" + str(initial_cost_inverter))
    print("Cost of charge controllers: $" + str(initial_cost_controller))
    

    
    # #Inital upfront cost
    # initial_cost = (initial_cost_solar + battery_initial_cost + initial_cost_inverter + initial_cost_controller)
    # print("Initial cost for system: $" + str(initial_cost))
    
    #OPERATION COSTS
    maintenance_battery = battery_initial_cost * 0.01               #Battery maintenance is 1% of inital battery cost
    maintenance_inverter = inverter_maintenance * inverter_count
    maintenance_controller = controller_maintenance * controller_count
    
    # maintenance_per_year = (panel_maintenance + maintenance_battery + maintenance_inverter + maintenance_controller)
    # print("Yearly Maintenance: $" + str(maintenance_per_year))
    # print()
    # print("Total cost: $" + str((maintenance_per_year*25) + initial_cost))
#*****************************************************************************

    #Cost of generator
    solar_life_cost = 0
    rate = 0.0231     #This is the 20 year average rate of inflation
    year_energy = ((sum(lowest_solar_value)) * 365)
    print()
    
    print("Yearly operating costs:")
    for i in range(0,21):
        #cost to operate generator in year i
        solar_year_cost = (panel_maintenance * panel_count) + maintenance_battery + maintenance_inverter + maintenance_controller
        
        if i == 0:  #inital investment
            solar_year_cost += (battery_initial_cost + initial_cost_solar + initial_cost_inverter + initial_cost_controller)    
        if i == 10: 
            solar_year_cost += battery_initial_cost - (battery_initial_cost * 0.1)        #cost of new batteries - resale value of old batteries
        # if i == 20:
        #     solar_year_cost += battery_initial_cost - (battery_initial_cost * 0.1) 
        
        
        #Adjust yearly cost for inflation
        solar_cost_inflation = inflation(solar_year_cost, rate, i)
        LCOE_inflation = inflation(LCOE_solar, rate, i)
        new_way = (LCOE_inflation * year_energy) + solar_cost_inflation
        print("Cost to operate in year " + str(i) + " is $" + str('{0:.2f}'.format(solar_cost_inflation)))       
        #then save to spreadsheet
        sheet.cell(row = i+2, column = 2).value = new_way
        solar_life_cost += new_way
        sheet.cell(row = i+2, column = 3).value = solar_life_cost

        #add to running sum
        
    
    print()
    print("Cost of this solar array over 20 years is: $" + str('{0:.2f}'.format(solar_life_cost)))
    #THIS METHOD IS NOT SUPER ACCURAGE. 
    #print("Total energy used over lifetime " + str('{0:.2f}'.format(life_energy)) + str(" kW"))
    #Determine the LOCE 
    #life_energy = ((sum(lowest_solar_value)) * 365 * 20)
    print('life Energy')
    print(life_energy)
    print("solar_life_cost:")
    print(solar_life_cost)
    LCOE_solar_2 = (solar_life_cost / life_energy)
    LCOE_solar_3 = LCOE_solar_2 + LCOE_solar
    print("The LOCE of solar before adding is: " + str('{0:.3f}'.format(LCOE_solar_3)) + str(" $/kWh"))




def main():
    
    build_spreadsheet()
    avaliable_solar()
    spreadsheet = wb.save('Solar_Cost_Round_2.csv')

    return

if __name__ == "__main__":
    main()