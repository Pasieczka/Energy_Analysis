"""
This script will analyze the cost of only diesel generators.
There will be 3 Cummins DSGAC-150 gensets used in this example. 

"""


import numpy as np
from numpy import matrix
from scipy.optimize import minimize
import math
import openpyxl
import csv
import sys
import pandas as pd


#Generator Spec             UNIT
#CONSTANTS
diesel_cost = 1              #$/L
inflation_rate = 0.0261
fuel_per_day = 1083.60            #L/hour
gen_operting_cost = diesel_cost * fuel_per_day * 365  #operating costs for 1 year
gen_replacement_cost = 0.0004 * 135          #$/W


gen_maintenance = 5000      #$/year
num_gen = 3                 #using 3 generators
#life_energy = 287.24 * 365 * 24 *20
life_energy = 2301873.1 * 20        #kW/ year * 20 years #total energy used in a year span multiplied by 20 years


#load in kW for January 30th 1990 (Highest daily average load)
load_data = [278.4 ,276.2, 274.1, 271, 269.3, 279.7, 311.8, 299.3, 302.3, 313.2, 321.3, 322.5, 345, 336.7, 329.1, 325.5, 315.9, 318.3, 328.1, 364.6, 339.9, 315.6, 308.1, 288.1]


#Defining workbook we will save values to 
wb = openpyxl.Workbook()
sheet = wb.active



def inflation(present_value, rate, time):
    
    future_value = ((present_value) * ((1.0 + (rate))**(time)))
    
    return future_value




def cost_of_generators():
    gen_life_cost = 0
    rate = 0.0231     #This is 20 year average rate of inflation
    for i in range(1,21):
        
        #cost to operate generator in year i
        gen_year_cost = (gen_maintenance * num_gen ) + gen_operting_cost
        if i == 5 or i == 10 or i == 15:
            gen_year_cost += gen_replacement_cost * num_gen
        
        #Then run through inflation cost
        gen_cost_inflation = inflation(gen_year_cost, rate, i)
        print("Cost to operate in year " + str(i) + " is $" + str('{0:.2f}'.format(gen_cost_inflation)))
      
        
        #then save to spreadsheet
        sheet.cell(row = i+1, column = 2).value = gen_cost_inflation
        #add to running sum
        gen_life_cost += gen_cost_inflation
        sheet.cell(row = i+1, column = 3).value = gen_life_cost

        
    print()
    print("Cost of generators over 20 years is: $" + str('{0:.2f}'.format(gen_life_cost)))
    
    #THIS METHOD IS NOT SUPER ACCURAGE. 
    print("Total energy used over lifetime " + str('{0:.2f}'.format(life_energy)) + str(" kW"))
    
    #Determine the LOCE 
    LOCE_generator = (gen_life_cost / life_energy)
    print("The LOCE of 3 diesel generators is: " + str('{0:.3f}'.format(LOCE_generator)) + str(" $/kWh"))
    
    return



def build_spreadsheet():
    
    #Defining columns and titles
    sheet.cell(row = 1, column = 1).value = "Year"
    sheet.cell(row = 1, column = 2).value = "Cost"
    sheet.cell(row = 1, column = 3).value = "Total Cost"


    for j in range (2,22):
        sheet.cell(row = j, column = 1).value = (j-1)
        
    return




def main():
    build_spreadsheet()
    cost_of_generators()
    spreadsheet = wb.save('Generator_Cost_Round_2.csv')

    return

if __name__ == "__main__":
    main()
