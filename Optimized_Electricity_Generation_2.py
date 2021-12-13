# -*- coding: utf-8 -*-
"""
The purpose of this scrip to find the optimal number of solar panels for the lowest cost production of energy.
This script will output a .csv file which can then be analyzed for the optimal balance in a hybrid system.
"""
#COST OF TRIAL

import numpy as np
from numpy import matrix
from scipy.optimize import minimize
import math
import openpyxl
import csv
import sys
import pandas as pd

import matplotlib.pyplot as plt
from itertools import combinations

#CONSTNATS

#Battery Spec               UNIT
battery_capacity = 13.5     #kWh
battery_cost = 8500         #$
battery_maintenance = 50    #$/year
battery_dc_voltage = 50     #V


#Solar Panel Specs
panel_cost = 220            #$
panel_maintenance = 10      #$/year
panel_wattage = 0.425         #kW
efficiency = 0.90

#Inverter Specs
inverter_cost = 0           #$ #set to zero because battery: Tesla Powerwall 2 has inverter combined. 
inverter_maintenance = 0    #$/year

#Charge Controller Specs
controller_cost =15         #$
controller_maintenance = 5    #$/year #accounts for having to replace controllers
controller_amperage = 30 

#Generator Spec             UNIT
diesel_cost = 1             #$/L
inflation_rate = 0.0261     #%
#fuel_per_day = 1083.60      #L/hour
#gen_operting_cost = diesel_cost * fuel_per_day * 365  #operating costs for 1 year
gen_replacement_cost = 0.0004 * 135          #$/W


gen_maintenance = 5000      #$/year
num_gen = 3                 #using 3 generators
#BUFFER
buffer = 1.2255 #Use this buffer to play around with what level of factor you want for the array
#peak load is 376.1 kW (from Nov 20th @ 19:00) / 306.875 (daily average from peak day) = 1.2255
rate = 0.0231     #This is 20 year average rate of inflation
#DATA

#Load data from January 30th 1990 (kW)
load_data = [278.4 ,276.2, 274.1, 271, 269.3, 279.7, 311.8, 299.3, 302.3, 313.2, 321.3, 322.5, 345, 336.7, 329.1, 325.5, 315.9, 318.3, 328.1, 364.6, 339.9, 315.6, 308.1, 288.1]
load_data = [element * buffer for element in load_data]
trial3 = {}

#Solar Data from June 9th 2014 (hourly average)
solar_data = [0, 0, 0, 0, 0 , 0.675, 2.883333333, 5.2, 6.966666667, 9.35, 10.775, 10.50833333, 9.366666667, 7.1, 3.95, 1.108333333, 0, 0, 0, 0, 0, 0, 0, 0] #June 9th
#load_data = [278.4 ,276.2, 274.1, 271, 269.3, 279.7, 311.8, 299.3, 302.3, 313.2, 321.3, 322.5, 345, 336.7, 329.1, 325.5, 315.9, 318.3, 328.1, 364.6, 339.9, 315.6, 308.1, 288.1]

    
def generator_cost(generator_output, p, k):
    
    gen_year_cost = 0
   ### print("GENERATOR COST FUNCTION")
    #print(generator_output)
    day_fuel_rate = 0
    for hour in generator_output:
        ##print(hour)
        #gen_fuel_rate = fuel_rate(float(hour))
        #day_fuel_rate += gen_fuel_rate
        gen_fuel_rate = 1083 *(1/k)    #fuel used per day #1083 is maximum fuel consumption for just generators
                              
        #I was unable to get the function fuel_rate() to work in this hybrid model. I will therefore represent
        #the fuel consumptio as a ratio of k. 135 being zero. 1 being 1083 l/hr. 
    ##print("daily fuel consumption")
    #print(day_fuel_rate)
        
    gen_operting_cost = (gen_fuel_rate *365)
    #cost to operate generator in year i
    gen_year_cost = (gen_maintenance * num_gen ) + gen_operting_cost
    if p == 5 or p == 10 or p == 15: #Check if we need to replace generators this year
        gen_year_cost += gen_replacement_cost * num_gen
     
    total_gen_cost_inflation = inflation(gen_year_cost, rate, p) #adjust price for inflation
    return total_gen_cost_inflation


def solar_panel_cost(panel_count, p):
    if p == 1: #inital cost of solar panels
        total_solar_cost = (panel_cost + panel_maintenance) * panel_count
    else: #Every other year p
        total_solar_cost = (panel_maintenance) * panel_count

    total_solar_cost_inflation = inflation(total_solar_cost, rate, p)
    return total_solar_cost_inflation

def total_controller_cost(controller_count, p):
    if p == 1: #inital cost of controller
        controller_cost_year = controller_count * (controller_cost + controller_maintenance )
    else: 
        controller_cost_year = controller_count * (controller_maintenance )

    controller_cost_year_inflation = inflation(controller_cost_year, rate, p)
    return controller_cost_year_inflation

#This is determining the cost of batteries for year p
def battery_cost(battery_hours, battery_storage, p):
    battery_per_hour_load = battery_storage / battery_hours
    battery_initial_cost = 0 
    
    #determining the cost of the battery based off the size
    if battery_hours > 0:
        for k in range(battery_hours):   #this was an idea to step down battery hours until it's zero
            if k == 0:
                battery_initial_cost += battery_per_hour_load * 500   #cost to install batteries for the first 2 hours of a system
            if k == 1: 
                battery_initial_cost += battery_per_hour_load * 500   #cost to install batteries for the first 2 hours of a system
            if k >= 2:
                battery_initial_cost += battery_per_hour_load * 25   #cost to install batteries for the first 2 hours of a system
   
    #intalling new battery
    if p == 0:  #inital investment
        battery_initial_cost += (battery_initial_cost )    
    if p == 10: 
        battery_initial_cost += battery_initial_cost - (battery_initial_cost * 0.1)        #cost of new batteries - resale value of old batteries

   
    battery_inital_cost_inflation = inflation(battery_initial_cost, rate, p)
    #print('{0:.2f}'.format(battery_initial_cost))
    #print("range of battery_hours")
    # print("INITIAL COST")
    # print(battery_initial_cost)
    
    return battery_inital_cost_inflation





#START OF SOMETHING GREAT
#load in kW for January 30th 1990 (Highest daily average load)

def inflation(present_value, rate, time):
    future_value = ((present_value) * ((1.0 + (rate))**(time)))
    return future_value



def build_spreadsheet(sheet):
    
    #Defining columns and titles
    sheet.cell(row = 1, column = 1).value = "Year"
    sheet.cell(row = 1, column = 2).value = "Cost"
    sheet.cell(row = 1, column = 3).value = "Total Cost"

    for j in range (1,22):  #creating years
        sheet.cell(row = j+1, column = 1).value = (j-1)
    return

def fuel_rate(hour):
    

    #OPTIMIZATION COST FOR ONLY GENERATORs
    #Code adapted from: https://www.geeksforgeeks.org/combinations-with-repetitions/
    #This function shows the fuel consumption rate of "n" generators.
    #The output will show the total fuel rate for n generators, follwed by each individual generator fuel rate
    
    #Combination code for generator
    # Gather all combinations of size r in an array of size n
     
    
    #Cost of fuel
    fuel_cost = 1           #$/L
    
    
    #Fuel consumption rate (L/hr)
    #100_rate = 40.9
    # 75_rate = 33.2
    # 50_rate = 23.9
    # 25_rate = 14.7
    
    #Setting parameters
    arr = [ 135, 101.25, 67.5, 33.75 ] #different power outputs for generator
    r = 3    #possible number of generators 
    n = len(arr) - 1
    
    
    def CombinationRepetitionUtil(chosen, arr, index, r, start, end, hour, trial3):
        # load_data = [278.4 ,276.2, 274.1, 271, 269.3, 279.7, 311.8, 299.3, 302.3, 313.2, 321.3, 322.5, 345, 336.7, 329.1, 325.5, 315.9, 318.3, 328.1, 364.6, 339.9, 315.6, 308.1, 288.1]
        # load_data = [element * buffer for element in load_data]
        if index == r:
            range_sum = 0
            trial_rate = 0
    
            for j in range(r):
               
                if chosen[j] == 135:
                    trial_rate = trial_rate + (40.9)
                if chosen[j] == 101.25:
                    trial_rate = trial_rate + (33.2)
                if chosen[j] == 67.5:
                    trial_rate = trial_rate + (23.9)
                if chosen[j] == 33.75:
                    trial_rate = trial_rate + (14.7)  
                range_sum += chosen[j]
           
            if range_sum >= hour: 
                trial3[range_sum] = float('{0:.2f}'.format(trial_rate))   
            return 
        # When there are no more elements to put in chosen{}
        if start > n:
            return 
    
        chosen[index] = arr[start]
        CombinationRepetitionUtil(chosen, arr, index + 1, r, start, end, hour, trial3)
        CombinationRepetitionUtil(chosen, arr, index, r, start + 1, end, hour, trial3)
     
  
    def CombinationRepetition(arr, n, r):
        
        # A temporary array to store all combination one by one
        chosen = [0] * r
        #trial3 = {}
        fuel_used = 0
    
        #for load in hourly:
        CombinationRepetitionUtil(chosen, arr, 0, r, 0, n, hour, trial3)
        #print(trial)
        min_rate = min(trial3.items())
        fuel_used += min_rate[1]
        #trial = {}
            
        # print()
        # print("TOTAL fuel used in day")
        # print(('{0:.2f}'.format(fuel_used)) + " L/day")
        
        return fuel_used
    fuel_used = CombinationRepetition(arr, n, r)
    
    return float('{0:.2f}'.format(fuel_used))



def whole_number(battery_count):
    battery_count_int = battery_count.is_integer()
    if (battery_count_int) == False:
        battery_count += 1
        battery_count = math.trunc(battery_count)
    return battery_count








#This function will determine the lowest cost for the load
def optimize():
    print("in optimize")
    lowest_multiplier = 1000
    biggest_price = 1000000000
    
    for k in range(135, 27, -1):   #we know multiplying the solar data by 135 will power the entire array. It could be any arbitrary number. 
        print()
        print("Multiplying factor is: " + str(k))    
        solar_data_multiplied = [element * k for element in solar_data]
        battery_hours = 0 #number of hours the battery will be used
        hours = 0
        battery_storage_need = 0
        
        #finding difference between matrices 
        load_data_matrix = np.array(load_data)
        solar_data_matrix = np.array(solar_data_multiplied)
        difference2 = solar_data_multiplied - load_data_matrix
        difference = solar_data_matrix - load_data_matrix 
        #*****************
        #Finding battery storage
        battery_storage = 0
        battery_hours = 0
        for entry in difference:
            if entry < 0:
                battery_storage += entry
            if entry > 0:
                battery_hours += 1
                
        #************** NEW TRY
        battery_charging = []
        #for load in load_data_matrix:
            #list(np.array(a) - np.array(b))
        battery_charging = np.array(solar_data_multiplied) - np.array(load_data)
        battery_charging_sum = sum(battery_charging)
        print(battery_charging_sum)
        battery_offset_per_hour = battery_charging / 12   
        print('BATTERY CHAARGING')
        #print(battery_offset_per_hour)
        generator_energy_demand = np.array(load_data) - np.array(battery_offset_per_hour) - np.array(solar_data_multiplied)
        generator_output = []
        generator_off = 0
        
        for gen in generator_energy_demand:
            if gen > 0:
                generator_output.append(gen)
            else:
                generator_output.append(0)
                generator_off +=1
        #print(generator_output)
        generator_hour = (sum(generator_output)/ (24-generator_off))
        
        
        
        battery_storage = abs(battery_storage)
        print("Battery storage required: " + str('{0:.2f}'.format(battery_charging_sum)) + (" kW"))
        #*****************
        
        #getting battery figures
        for data in difference:
            if data > 0:
                battery_storage_need += data
                hours += 1
           
        
        
        battery_hours = (24 - hours)
        battery_per_hour = battery_storage_need / (24-hours) #How much the battery will use per hour
        print("size of grid: " + str(solar_data_multiplied[10]) + " kW")
        print("QUANTITIES:")
        print("HOURS to store energy in battery: " + str(hours))
        print("Battery load per hour: " + str('{0:.2f}'.format(battery_per_hour)) + " kWh")
        print("Battery storage: " + str('{0:.2f}'.format(battery_storage_need)) + " kW")

        #generator 
        generator_power = sum(load_data) - sum(solar_data_multiplied) - battery_storage_need
        generator_power_2 = abs(sum(solar_data_multiplied) - battery_storage_need - sum(load_data))
        print()
        print("GENERATOR_POWER")
        
        print("generator_power: " + str('{0:.2f}'.format(generator_power_2)) + " kW")
        #print(generator_power/battery_hours)
        generator_hourly_rate = abs(generator_power_2)/battery_hours
        
        
        #Number of Solar Panels
        pv_array = ((solar_data_multiplied[10])) #Maximum kW required to be output at the peak hour for this multiplying factor
        panel_count = pv_array / panel_wattage
        panel_count = whole_number(panel_count)
        print("Number of solar panels required: " + str(panel_count))
                
        #Number of controllers
        controller_size = ((panel_count * (panel_wattage * 1000)) / battery_dc_voltage) #amp
        controller_count = controller_size / (controller_amperage)
        controller_count = whole_number(controller_count)
        print("Number of controllers required: " + str(controller_count))

                
            #generator_demand = load_data_matrix -
        
        
        #generator_load = 
        #battery_load = 
        
        
        #COST
        trial_running_sum = 0
        for p in range(1,21):  #This will step through each year
       #    trial_cost = cost_generator(p) + solar_panel_cost(panel_count, p) + cost_battery(battery_hours, battery_storage_need, p) + cost_controller()
           #trial_cost = solar_panel_cost(panel_count, p) + battery_cost(battery_hours, battery_storage_need, p) + controller_cost(controller_count, p)
           p = float(p)
           
           panel_trial_cost = solar_panel_cost(panel_count, p)
           battery_trial_cost = battery_cost(battery_hours, battery_storage_need, p)
           controller_trial_cost = total_controller_cost(controller_count, p)
           generator_trial_cost = generator_cost(generator_output, p, k)
           trial_cost = (panel_trial_cost + battery_trial_cost + controller_trial_cost + generator_trial_cost)

           trial_running_sum += trial_cost
           
           lowest_cost = []
           lowest_cost.append('{0:.2f}'.format(trial_running_sum))
        print("Trial TOTAL COST: $" + str('{0:.2f}'.format(trial_running_sum)))
        trial3[k] = float('{0:.2f}'.format(trial_running_sum)) 
        
        
        
       #ADD TRIAL COST TO A SPREADHSHEET
           # sheet.cell(row = p+1, column = K+1).value = trial_cost

        #if trial_cost < biggest_price:
        #     lowest_multiplier = n
           # sheet.cell(row = i+2, column = 3).value = solar_life_cost #Add system cost
    print()
    print("Lowest cost")
    #print((trial3.items()))
    #print lowest_multiplier
        #print all the other costs and values associated with it here. 
        
    print((lowest_cost))
    return

def main():
    #Defining workbook we will save values to 
    wb = openpyxl.Workbook()
    sheet = wb.active
    build_spreadsheet(sheet)
    
    optimize()
   
    
    spreadsheet = wb.save('Hybrid_Cost_Round_2.csv')

    return


if __name__ == "__main__":
    main()
    
    
# Send to a csv
# •	Column 1 = number of panels
# •	Column 2 = cost over 25 years
# Data to use 
# •	Lowest solar production day (June 9th)
# •	Highest load day

